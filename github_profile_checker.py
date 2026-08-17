#!/usr/bin/env python3
"""
GitHub Batch Profile Checker
-----------------------------
For DevRel / HR use: pulls public profile stats for a list of GitHub
usernames — account age, bio, followers, public repo count, total stars,
top languages, and an estimated recent commit count (via the Events API,
last 90 days of public activity).

Usage:
    python github_profile_checker.py

You'll be prompted for:
    1. A GitHub Personal Access Token (PAT) - optional but strongly
       recommended, since it raises your rate limit from 60 to 5000
       requests/hour. Create one at https://github.com/settings/tokens
       (no special scopes needed for public data - "no scopes" is fine).
    2. A list of GitHub usernames (comma or newline separated), or a
       path to a .txt file with one username per line.

Output:
    - Printed summary table in the terminal
    - You'll be asked which file format(s) to save: CSV, XLSX, TXT, or any
      combination. Files are named github_profile_report.csv / .xlsx / .txt

Notes on limitations:
    - GitHub's API does not expose a user's *total* lifetime commit count
      directly (that's not something the platform surfaces, even to the
      user themselves, since it would require scanning every repo they've
      ever touched, including private ones you don't have access to).
      This script instead reports:
        * public_repos: number of public repositories
        * total_stars: stars across their public repos
        * recent_public_events_90d: count of public activity events
          (pushes, PRs, issues, etc.) in roughly the last 90 days
        * recent_push_events_90d: subset of the above that are code pushes
      These are reasonable, honest proxies for public activity level -
      treat them as such rather than as an exact commit count.
"""

import csv
import os
import sys
import time
from datetime import datetime, timezone
import getpass

try:
    import requests
except ImportError:
    print("This script needs the 'requests' library. Install it with:")
    print("    pip install requests")
    sys.exit(1)

FIELDNAMES = [
    "username", "name", "bio", "company", "location",
    "account_created", "account_age_years", "public_repos", "public_gists",
    "followers", "following", "total_stars_owned_repos", "top_languages",
    "recent_public_events_90d", "recent_push_events_90d", "hireable",
    "profile_url", "error",
]

API_ROOT = "https://api.github.com"


def get_headers(token):
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def check_rate_limit(headers):
    r = requests.get(f"{API_ROOT}/rate_limit", headers=headers)
    if r.status_code == 200:
        core = r.json()["resources"]["core"]
        print(f"Rate limit: {core['remaining']}/{core['limit']} requests remaining.\n")


def fetch_user(username, headers):
    r = requests.get(f"{API_ROOT}/users/{username}", headers=headers)
    if r.status_code == 404:
        return {"username": username, "error": "User not found"}
    if r.status_code == 403:
        return {"username": username, "error": "Rate limited / forbidden - check your token"}
    if r.status_code != 200:
        return {"username": username, "error": f"HTTP {r.status_code}"}
    return r.json()


def fetch_repo_stats(username, headers, max_pages=5):
    """Sum stars, and collect top languages, across the user's public repos."""
    total_stars = 0
    languages = {}
    page = 1
    while page <= max_pages:
        r = requests.get(
            f"{API_ROOT}/users/{username}/repos",
            headers=headers,
            params={"per_page": 100, "page": page, "type": "owner"},
        )
        if r.status_code != 200:
            break
        repos = r.json()
        if not repos:
            break
        for repo in repos:
            total_stars += repo.get("stargazers_count", 0) or 0
            lang = repo.get("language")
            if lang:
                languages[lang] = languages.get(lang, 0) + 1
        if len(repos) < 100:
            break
        page += 1
    top_languages = sorted(languages.items(), key=lambda x: -x[1])[:3]
    top_languages_str = ", ".join(f"{lang} ({count})" for lang, count in top_languages)
    return total_stars, top_languages_str


def fetch_recent_activity(username, headers, days=90):
    """Estimate recent public activity from the Events API (max ~last 90-300 events, GitHub-limited)."""
    cutoff = datetime.now(timezone.utc).timestamp() - days * 86400
    total_events = 0
    push_events = 0
    page = 1
    while page <= 3:  # GitHub only returns the most recent ~300 events anyway
        r = requests.get(
            f"{API_ROOT}/users/{username}/events/public",
            headers=headers,
            params={"per_page": 100, "page": page},
        )
        if r.status_code != 200:
            break
        events = r.json()
        if not events:
            break
        stop = False
        for ev in events:
            ev_time = datetime.strptime(ev["created_at"], "%Y-%m-%dT%H:%M:%SZ").replace(
                tzinfo=timezone.utc
            ).timestamp()
            if ev_time < cutoff:
                stop = True
                break
            total_events += 1
            if ev["type"] == "PushEvent":
                push_events += 1
        if stop or len(events) < 100:
            break
        page += 1
    return total_events, push_events


def account_age_years(created_at_str):
    created = datetime.strptime(created_at_str, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
    now = datetime.now(timezone.utc)
    return round((now - created).days / 365.25, 1)


def get_usernames_input():
    print("\nEnter GitHub usernames (comma-separated), OR a path to a .txt file")
    print("with one username per line, then press Enter:")
    raw = input("> ").strip()

    if raw.lower().endswith(".txt"):
        try:
            with open(raw, "r") as f:
                usernames = [
                    line.strip() for line in f
                    if line.strip() and not line.strip().startswith("#")
                ]
        except FileNotFoundError:
            print(f"File not found: {raw}")
            sys.exit(1)
    else:
        usernames = [u.strip() for u in raw.replace("\n", ",").split(",") if u.strip()]

    # de-dupe, preserve order
    seen = set()
    deduped = []
    for u in usernames:
        if u.lower() not in seen:
            seen.add(u.lower())
            deduped.append(u)
    return deduped


def get_output_formats():
    print("\nWhich format(s) should the report be saved in?")
    print("  1) CSV")
    print("  2) XLSX (Excel)")
    print("  3) TXT (plain text, aligned columns)")
    print("  4) All of the above")
    raw = input("Enter numbers separated by commas (e.g. 1,2) [default: 1]: ").strip()
    if not raw:
        return {"csv"}
    choice_map = {"1": "csv", "2": "xlsx", "3": "txt", "4": "all"}
    chosen = set()
    for part in raw.split(","):
        part = part.strip()
        fmt = choice_map.get(part)
        if fmt == "all":
            return {"csv", "xlsx", "txt"}
        elif fmt:
            chosen.add(fmt)
    return chosen or {"csv"}


def write_csv(results, out_file="github_profile_report.csv"):
    with open(out_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        for row in results:
            writer.writerow({k: row.get(k, "") for k in FIELDNAMES})
    print(f"Saved: {out_file}")


def write_xlsx(results, out_file="github_profile_report.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("XLSX output needs 'openpyxl'. Install it with: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "GitHub Profiles"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="24292F", end_color="24292F", fill_type="solid")
    body_font = Font(name="Arial")

    for col_idx, name in enumerate(FIELDNAMES, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(results, start=2):
        for col_idx, key in enumerate(FIELDNAMES, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.font = body_font

    # reasonable column widths
    widths = {
        "username": 18, "name": 20, "bio": 40, "company": 18, "location": 18,
        "account_created": 14, "account_age_years": 14, "public_repos": 12,
        "public_gists": 12, "followers": 11, "following": 11,
        "total_stars_owned_repos": 16, "top_languages": 28,
        "recent_public_events_90d": 16, "recent_push_events_90d": 16,
        "hireable": 10, "profile_url": 35, "error": 20,
    }
    for col_idx, name in enumerate(FIELDNAMES, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 14)

    ws.freeze_panes = "A2"
    wb.save(out_file)
    print(f"Saved: {out_file}")


def write_txt(results, out_file="github_profile_report.txt"):
    display_fields = [
        "username", "account_age_years", "public_repos",
        "total_stars_owned_repos", "followers", "recent_public_events_90d",
        "recent_push_events_90d", "top_languages", "profile_url", "error",
    ]
    col_widths = {f: max(len(f), 10) for f in display_fields}
    for row in results:
        for f in display_fields:
            col_widths[f] = max(col_widths[f], len(str(row.get(f, ""))))

    lines = []
    header = "  ".join(f.replace("_", " ").title().ljust(col_widths[f]) for f in display_fields)
    lines.append(header)
    lines.append("-" * len(header))
    for row in results:
        line = "  ".join(str(row.get(f, "")).ljust(col_widths[f]) for f in display_fields)
        lines.append(line)

    with open(out_file, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"Saved: {out_file}")


def main():
    print("=== GitHub Batch Profile Checker ===\n")

    # Allow non-interactive use (e.g. CI, cron, scripting) via env var
    token = os.environ.get("GITHUB_TOKEN", "").strip()
    if token:
        print("Using GITHUB_TOKEN from environment.")
    else:
        token = getpass.getpass(
            "GitHub Personal Access Token (input hidden, press Enter to skip - "
            "60 req/hr unauthenticated limit applies without it): "
        ).strip()

    headers = get_headers(token)
    check_rate_limit(headers)

    usernames = get_usernames_input()
    if not usernames:
        print("No usernames provided. Exiting.")
        sys.exit(0)

    results = []
    print(f"\nChecking {len(usernames)} user(s)...\n")

    for i, username in enumerate(usernames, 1):
        print(f"[{i}/{len(usernames)}] {username} ...", end=" ", flush=True)
        user = fetch_user(username, headers)

        if "error" in user:
            print(user["error"])
            results.append({"username": username, "error": user["error"]})
            continue

        total_stars, top_langs = fetch_repo_stats(username, headers)
        recent_events, recent_pushes = fetch_recent_activity(username, headers)

        row = {
            "username": user.get("login", username),
            "name": user.get("name") or "",
            "bio": (user.get("bio") or "").replace("\n", " "),
            "company": user.get("company") or "",
            "location": user.get("location") or "",
            "account_created": user.get("created_at", "")[:10],
            "account_age_years": account_age_years(user["created_at"]) if user.get("created_at") else "",
            "public_repos": user.get("public_repos", 0),
            "public_gists": user.get("public_gists", 0),
            "followers": user.get("followers", 0),
            "following": user.get("following", 0),
            "total_stars_owned_repos": total_stars,
            "top_languages": top_langs,
            "recent_public_events_90d": recent_events,
            "recent_push_events_90d": recent_pushes,
            "hireable": user.get("hireable", ""),
            "profile_url": user.get("html_url", ""),
            "error": "",
        }
        results.append(row)
        print("done")
        time.sleep(0.2)  # small courtesy delay

    # Write output file(s) in the chosen format(s)
    formats = get_output_formats()
    print()
    if "csv" in formats:
        write_csv(results)
    if "xlsx" in formats:
        write_xlsx(results)
    if "txt" in formats:
        write_txt(results)
    print()

    # Print quick summary table
    print(f"{'Username':<20}{'Age(yrs)':<10}{'Repos':<8}{'Stars':<8}{'Followers':<11}{'Events(90d)':<12}")
    print("-" * 75)
    for row in results:
        if row.get("error"):
            print(f"{row['username']:<20}ERROR: {row['error']}")
            continue
        print(
            f"{row['username']:<20}"
            f"{str(row['account_age_years']):<10}"
            f"{str(row['public_repos']):<8}"
            f"{str(row['total_stars_owned_repos']):<8}"
            f"{str(row['followers']):<11}"
            f"{str(row['recent_public_events_90d']):<12}"
        )


if __name__ == "__main__":
    main()
